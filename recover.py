import math


from openpyxl import load_workbook
from tabulate import tabulate
import scipy

#Внешняя функция:
def differ(list1, list2):
    MSE = [(i + j) ** 2 for i, j in zip(list1, list2)]
    return sum(MSE)/len(MSE)

wb = load_workbook(r"technical_data2.xlsx")
sheet = wb.worksheets[0] # Атрибут для книги эксель
inj_amount = 0 # Количество нагнетательных скважин
prod_amount = 0 # Количество добывающих скважин
f = {} # Коэффициент взаимовлияния
tau = {} # Коэффициент временного лага
qliq0 = {} # Начальные дебиты
inj_wells_rates = {} # Словарь ключи имена скважин, значения данные по закачке # TODO: Можно оставить как локальну переменную в методе
prod_wells_rates = {} # Данные по добыче # TODO: Можно  оставить как локальную переменную в методе
prod_CRM = {} # Об
MSE = {}
sum_prod_CRM = {} # Сумма всех CRM моделей (равна количеству добывающих скважин)
time = [] # Время, в целом оно не нужно - есть количество строк, то есть шагов "К"
print(qliq0)

def amount_of_params(sheet, inj_amount, prod_amount, prod_wells_rates, inj_wells_rates): # TODO: Мне кажется всю эту функцию можно оставить в init
    for i in range(2, sheet.max_row):
        time.append(sheet[i][0].value)
    for column in range(1, sheet.max_column):
        q_data = [] #временная переменная для добавления расходов доб/нагн скважин
        # Идем по листу и добавляем информацию в словари
        if 'INJ' in sheet[1][column].value:
            inj_amount += 1
            for i in range(2, sheet.max_row):
                q_data.append(sheet[i][column].value)
            inj_wells_rates.update({sheet[1][column].value: q_data})
        else:
            prod_amount += 1
            for i in range(2, sheet.max_row):
                q_data.append(sheet[i][column].value)
            prod_wells_rates.update({sheet[1][column].value: q_data})
            MSE.update({'MSE ' + sheet[1][column].value: None})
    # Создаем словари с нужными параметрами и названиями CRM-скважин-моделей
    for i in prod_wells_rates:
        for j in inj_wells_rates:
            f.update({'f_' + str(i[5::]) +  ' ' + str(j[4::]): None})
            tau.update({'tau_' + str(i[5::]) + ' ' + str(j[4::]): None})
            qliq0.update({'qliq0_' + str(i[5::]) +  ' ' + str(j[4::]): None})
            prod_CRM.update({i[5::] + ' & ' + j[4::]: []})
    return f

prod_CRM = amount_of_params(sheet, inj_amount, prod_amount, prod_wells_rates, inj_wells_rates)


for i in prod_CRM:
    print(i)

print(f)
print(qliq0)