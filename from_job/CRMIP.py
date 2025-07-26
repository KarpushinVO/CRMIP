import math
import scipy
import datetime as dt
import matplotlib.pyplot as plt
import matplotlib as mpl
from openpyxl import load_workbook
from copy import deepcopy

# Устанавливаем глобальный размер шрифта
mpl.rcParams['font.size'] = 7  # Задаем размер шрифта 10

# Внешняя функция:
def differ(list1, list2):
    """
    :param list1: список с int либо float
    :param list2: список с int либо float
    :return: возвращает среднеквадратическую ошибку
    """
    MSE = [(i + j) ** 2 for i, j in zip(list1, list2)]
    return sum(MSE) / len(MSE)

class ProxyModel:

    def __init__(self, excel):
        self.sheet = excel.worksheets[0]  # Атрибут для книги эксель
        self.inj_amount = 0  # Количество нагнетательных скважин
        self.prod_amount = 0  # Количество добывающих скважин
        self.f = {}  # Коэффициент взаимовлияния
        self.tau = {}  # Коэффициент временного лага
        self.qliq0 = {}  # Начальные дебиты
        self.inj_wells_rates = {}  # Словарь ключи имена скважин, значения данные по закачке # TODO: Можно оставить
        # как локальную переменную в методе
        self.prod_wells_rates = {}  # Данные по добыче # TODO: Можно оставить как локальную переменную в методе
        self.prod_CRM = {}  # Об
        self.MSE = {}
        self.sum_prod_CRM = {}  # Сумма всех CRM моделей (равна количеству добывающих скважин)
        self.time = []  # Время, в целом оно не нужно - есть количество строк, то есть шагов "К"

        for i in range(2, self.sheet.max_row):
            self.time.append(self.sheet[i][0].value)
        for column in range(1, self.sheet.max_column):
            q_data = []  # временная переменная для добавления расходов доб/нагн скважин
            # Идем по листу и добавляем информацию в словари
            if 'INJ' in self.sheet[1][column].value:
                self.inj_amount += 1
                for i in range(2, self.sheet.max_row):
                    q_data.append(self.sheet[i][column].value)
                self.inj_wells_rates.update({self.sheet[1][column].value: q_data})
            else:
                self.prod_amount += 1
                for i in range(2, self.sheet.max_row):
                    q_data.append(self.sheet[i][column].value)
                self.prod_wells_rates.update({self.sheet[1][column].value: q_data})
                self.MSE.update({'MSE ' + self.sheet[1][column].value: None})
        # Создаем словари с нужными параметрами и названиями CRM-скважин-моделей
        for i in self.prod_wells_rates:
            for j in self.inj_wells_rates:
                self.f.update({'f_' + str(i[5::]) + ' ' + str(j[4::]): None})
                self.tau.update({'tau_' + str(i[5::]) + ' ' + str(j[4::]): None})
                self.qliq0.update({'qliq0_' + str(i[5::]) + '&' + str(j[4::]): None})
                self.prod_CRM.update({i[5::] + ' & ' + j[4::]: []})
        # Подготовка словаря для суммирования CRM моделей по одной скважине и
        self.sum_prod_CRM = self.sum_prod_CRM.fromkeys(list(self.prod_wells_rates.keys()), [])
        self.sum_prod_CRM = {key + ' CRM': value for key, value in self.sum_prod_CRM.items()}

        # Приведение к формату времени
        self.time = [dt.datetime.strptime(i, "%d.%m.%Y") for i in self.time]

    def make_initial_conditions(self):
        for i, j in zip(self.f, self.tau):  # Задаем начальные значения для f, tau
            self.f[i] = round(1 / self.inj_amount, 2)
            self.tau[j] = 10
        auto_ini_qliq0 = list(self.qliq0.keys())  # список ключей начальных дебитов
        step = len(
            auto_ini_qliq0) // self.prod_amount  # расчет шага количество начальных дебитов (модельных) / количество доб. скважи
        q_0 = []  # Временная переменная, чтобы передать первые значения из дебитов скважин и распределить их потом
        # по стартовым дебитам CRM скважин
        for i in self.prod_wells_rates:
            q_0.append(self.prod_wells_rates[i][0] // step)  # здесь начальный дебит добывающей скважины делится на
            # количество CRM моделей приходящих на нее

        boundary2 = step  # Чтобы перескакивать с одной группы CRM скважин на другую
        boundary1 = 0  # Чтобы заполнять двойками, тройками

        for i in range(0, len(q_0)):
            for j in range(boundary1, boundary2):
                auto_ini_qliq0[j] = q_0[i]
            boundary2 += step
            boundary1 += step
        self.qliq0 = {i: j for i, j in zip(list(self.qliq0.keys()), auto_ini_qliq0)}
        return auto_ini_qliq0

    def crm_calculate(self,
                      in_package):  # TODO: Нужно донастроить функциюю по аргументам, в циклах используются атрибуты класса, а нужно аргументы прим.: тау f
        # Добавляем данные в списки словаря из списка q аргумента функции
        list_length = len(self.qliq0)  # временная переменная для создания списка параметров на вход

        q_input = in_package[0:list_length]
        self.qliq0 = self.qliq0.fromkeys(list(self.qliq0.keys()), in_package[0:list_length])
        tau_input = in_package[list_length: list_length * 2]
        f_input = in_package[list_length * 2: list_length * 3]
        prod_CRM_output = deepcopy(self.prod_CRM)

        for i, j in enumerate(prod_CRM_output):  # TODO: обернуть в функцию, чтобы было красиво
            prod_CRM_output[j].append(q_input[i])

        for i, j in enumerate(self.tau):
            # Значение по ключу j равно значению из тау инпут
            self.tau[j] = tau_input[i]
        for i, j in enumerate(self.f):
            self.f[j] = f_input[i]

        for well_i, tau_i, f_i in zip(prod_CRM_output, self.tau, self.f):
            f_i_for_inj = f_i.split(
                ' ')  # Мне нужно на текущей итерации достать из ключа подстроку с указанием нагнетательной скважины!
            for i in range(1, len(self.time)):
                prod_CRM_output[well_i].append(
                    prod_CRM_output[well_i][i - 1] * math.exp(-(i - (i - 1)) / self.tau[tau_i]) + (
                                1 - math.exp(-(i - (i - 1)) / self.tau[tau_i])) * self.f[f_i] *
                    self.inj_wells_rates['INJ ' + f_i_for_inj[1]][i])

        step = len(list(prod_CRM_output.values())) // self.prod_amount
        boundary2 = step
        boundary1 = 0

        for i in self.sum_prod_CRM:
            self.sum_prod_CRM[i] = list(map(sum, zip(*list(prod_CRM_output.values())[boundary1:boundary2])))
            boundary2 += step
            boundary1 += step

        # Подсчет невязки
        for i, j, k in zip(self.MSE, self.sum_prod_CRM, self.prod_wells_rates):
            self.MSE[i] = round(differ(self.sum_prod_CRM[j], self.prod_wells_rates[k]), 3)
        return sum(self.MSE.values())


wb = load_workbook(r"C:\Users\Виктор\Desktop\CRMIP\input_data\technical_data2.xlsx")
model1 = ProxyModel(wb)

# print(model1.crm_calculate([350, 233, 900, 676, 1, 1, 1, 1, 0.25, 0.25, 0.25, 0.25]))

# Первое значение qliq, tau, f
x0 = [647, 220, 850, 800, 122, 122, 122, 122, 0.1, 0.9, 0.5, 0.5]

# Границы
bounds_for_all = (
    (0, 200), (0, 200), (0, 200), (0, 200), (0, 200), (0, 100), (0, 100), (0, 100), (0.025, 1),
    (0.025, 1), (0.025, 1), (0.025, 1))

cons = [{'type': 'eq', 'fun': lambda x: x[8] + x[9] - 1},
        {'type': 'eq', 'fun': lambda x: x[10] + x[11] - 1}]

result = scipy.optimize.minimize(model1.crm_calculate, x0, method='SLSQP', constraints=cons, bounds=bounds_for_all)

# print(result.x)
print(result.fun)

new_params = result.x
model1.crm_calculate(new_params)

CRM_names = list(model1.sum_prod_CRM.keys())  # Имена CRM скважин
FACT_names = list(model1.prod_wells_rates.keys())  # Имена фактических скважин
# print(model1.sum_prod_CRM)

# Данные для графиков
ydata1 = model1.sum_prod_CRM[CRM_names[0]]
ydata2 = model1.prod_wells_rates[FACT_names[0]]

ydata3 = model1.sum_prod_CRM[CRM_names[1]]
ydata4 = model1.prod_wells_rates[FACT_names[1]]

# plot the data
plt.figure(figsize=(10, 5))

# Создаем фигуру и оси
plt.subplot(1, 2, 1)
plt.plot(model1.time, ydata1, label=CRM_names[0])
plt.plot(model1.time, ydata2, label=f'{FACT_names[1]} Факт')
plt.xticks(rotation=45)
plt.title(f'Сравнение фактических и моделируемых дебитов скважины {FACT_names[0]}')
plt.xlabel('Дата')
plt.ylabel('Дебит Qж, м3/сут')
plt.legend()

plt.subplot(1, 2, 2)
plt.plot(model1.time, ydata3, label=CRM_names[1])
plt.plot(model1.time, ydata4, label=f'{FACT_names[1]} Факт')
plt.xticks(rotation=45)
plt.title(f'Сравнение фактических и моделируемых дебитов скважины {FACT_names[1]}')
plt.xlabel('Дата')
plt.ylabel('Дебит Qж, м3/сут')
plt.legend()

# Отображаем график
plt.tight_layout()  # Улучшает отображение, если метки перекрываются
plt.show()
